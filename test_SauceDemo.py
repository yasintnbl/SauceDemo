from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from Constants import globalConstants as gc


class TestSauce:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(gc.URL)

    def teardown_method(self):
        self.driver.quit()

    def waitForElement(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(
            ec.visibility_of_element_located((locator)))

    def getInValidData():
        excelFile = openpyxl.load_workbook("data/SauceData.xlsx")
        selectedSheet = excelFile["invalidLogin"]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)

        return data

    def getValidData():
        excelFile = openpyxl.load_workbook("data/SauceData.xlsx")
        selectedSheet = excelFile["validLogin"]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)

        return data


    @pytest.mark.parametrize("username,password", getInValidData())
    def test_InvalidLogin(self, username, password):
        self.waitForElement((By.ID, gc.usernameID))
        self.driver.find_element(By.ID, gc.usernameID).send_keys(username)
        self.waitForElement((By.ID, gc.passwordID))
        self.driver.find_element(By.ID, gc.passwordID).send_keys(password)
        self.waitForElement((By.ID, gc.loginButtonID))
        self.driver.find_element(By.ID, gc.loginButtonID).click()

    @pytest.mark.parametrize("username,password", getValidData())
    def test_ValidLogin(self, username, password):
        self.waitForElement((By.ID, gc.usernameID))
        self.driver.find_element(By.ID, gc.usernameID).send_keys(username)
        self.waitForElement((By.ID, gc.passwordID))
        self.driver.find_element(By.ID, gc.passwordID).send_keys(password)
        self.waitForElement((By.ID, gc.loginButtonID))
        self.driver.find_element(By.ID, gc.loginButtonID).click()

    @pytest.mark.parametrize("username,password", getValidData())
    def test_BuyProduct(self, username, password):
        self.waitForElement((By.ID, gc.usernameID))
        self.driver.find_element(By.ID, gc.usernameID).send_keys(username)
        self.waitForElement((By.ID, gc.passwordID))
        self.driver.find_element(By.ID, gc.passwordID).send_keys(password)
        self.waitForElement((By.ID, gc.loginButtonID))
        self.driver.find_element(By.ID, gc.loginButtonID).click()
        self.waitForElement((By.ID, gc.addToCartID))
        self.driver.find_element(By.ID, gc.addToCartID).click()
        self.waitForElement((By.ID, gc.shoppingButtonID))
        self.driver.find_element(By.ID, gc.shoppingButtonID).click()
        self.waitForElement((By.ID, gc.checkoutButtonID))
        self.driver.find_element(By.ID, gc.checkoutButtonID).click()
        self.waitForElement((By.ID, gc.firstNameID))
        self.driver.find_element(By.ID, gc.firstNameID).send_keys("Yasin")
        self.waitForElement((By.ID, gc.lastNameID))
        self.driver.find_element(By.ID, gc.lastNameID).send_keys("Tonbul")
        self.waitForElement((By.ID, gc.zipCodeID))
        self.driver.find_element(By.ID, gc.zipCodeID).send_keys("34692")
        self.waitForElement((By.ID,gc.continueButtonID))
        self.driver.find_element(By.ID,gc.continueButtonID).click()
        self.waitForElement((By.ID,gc.finishButtonID))
        self.driver.find_element(By.ID,gc.finishButtonID).click()

    @pytest.mark.parametrize("username,password", getValidData())
    def test_logOut(self, username, password):
        self.waitForElement((By.ID, gc.usernameID))
        self.driver.find_element(By.ID, gc.usernameID).send_keys(username)
        self.waitForElement((By.ID, gc.passwordID))
        self.driver.find_element(By.ID, gc.passwordID).send_keys(password)
        self.waitForElement((By.ID, gc.loginButtonID))
        self.driver.find_element(By.ID, gc.loginButtonID).click()
        self.waitForElement((By.ID, gc.menuButtonID))
        self.driver.find_element(By.ID, gc.menuButtonID).click()
        self.waitForElement((By.ID, gc.logoutButtonID))
        self.driver.find_element(By.ID, gc.logoutButtonID).click()

    
    @pytest.mark.parametrize("username,password", getValidData())
    def test_sortToPrice(self, username, password):
        self.waitForElement((By.ID, gc.usernameID))
        self.driver.find_element(By.ID, gc.usernameID).send_keys(username)
        self.waitForElement((By.ID, gc.passwordID))
        self.driver.find_element(By.ID, gc.passwordID).send_keys(password)
        self.waitForElement((By.ID, gc.loginButtonID))
        self.driver.find_element(By.ID, gc.loginButtonID).click()
        self.waitForElement((By.CLASS_NAME, gc.sortByButtonCS))
        self.driver.find_element(By.CLASS_NAME, gc.sortByButtonCS).click()
        self.waitForElement((By.XPATH, gc.lowToHighButtonXpath))
        self.driver.find_element(By.XPATH, gc.lowToHighButtonXpath).click()
        

       

